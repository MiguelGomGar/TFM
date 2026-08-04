"""Formatting helpers for publication-ready, multi-sheet Excel workbooks.

CSV cannot represent a multi-level header or row-group without repeating
labels on every line, which is unreadable once a table carries more than one
metric or partition. These helpers write each table's natural pandas
MultiIndex straight to a worksheet (openpyxl merges the repeated header/index
cells for free) and apply a single consistent visual style on top: bold dark
header, light banded rows, thin borders, frozen header/index panes and
auto-sized columns.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INDEX_FONT = Font(bold=True)
STRIPE_FILL = PatternFill("solid", fgColor="F3F4F6")
_BORDER_SIDE = Side(style="thin", color="D1D5DB")
CELL_BORDER = Border(
    left=_BORDER_SIDE, right=_BORDER_SIDE, top=_BORDER_SIDE, bottom=_BORDER_SIDE
)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")

MIN_COLUMN_WIDTH = 10
MAX_COLUMN_WIDTH = 42
NA_LABEL = "—"  # em dash, used in place of blank/NaN cells


def _style_worksheet(
    worksheet: Worksheet, n_header_rows: int, n_index_cols: int, n_data_rows: int
) -> None:
    """Apply the header/body/border/width style to a freshly written table.

    Parameters
    ----------
    worksheet : openpyxl.worksheet.worksheet.Worksheet
        Worksheet a DataFrame was just written to via ``DataFrame.to_excel``.
    n_header_rows : int
        Number of rows occupied by the column header (its MultiIndex depth,
        plus one extra row pandas inserts for the index name when the column
        header itself is a MultiIndex).
    n_index_cols : int
        Number of leading columns occupied by the row index (its MultiIndex
        depth).
    n_data_rows : int
        Number of data rows below the header.
    """
    max_row = n_header_rows + n_data_rows
    max_col = n_index_cols + (worksheet.max_column - n_index_cols)

    for row in worksheet.iter_rows(min_row=1, max_row=n_header_rows, max_col=max_col):
        for cell in row:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = CELL_BORDER

    for offset, row in enumerate(
        worksheet.iter_rows(
            min_row=n_header_rows + 1, max_row=max_row, max_col=max_col
        )
    ):
        stripe = offset % 2 == 1
        for col_idx, cell in enumerate(row, start=1):
            cell.border = CELL_BORDER
            if col_idx <= n_index_cols:
                cell.alignment = LEFT_ALIGNMENT
                cell.font = INDEX_FONT
            else:
                cell.alignment = CENTER_ALIGNMENT
            if stripe:
                cell.fill = STRIPE_FILL

    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        lengths = [
            len(str(cell.value))
            for cell in worksheet[letter][:max_row]
            if cell.value is not None
        ]
        width = max(lengths, default=MIN_COLUMN_WIDTH) + 2
        worksheet.column_dimensions[letter].width = min(
            max(width, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH
        )

    for row_idx in range(1, n_header_rows + 1):
        worksheet.row_dimensions[row_idx].height = 20

    worksheet.freeze_panes = worksheet.cell(
        row=n_header_rows + 1, column=n_index_cols + 1
    ).coordinate
    worksheet.sheet_view.showGridLines = False


def save_excel_workbook(
    sheets: dict[str, pd.DataFrame],
    file_path: str | Path,
    n_header_rows: int = 1,
    n_index_cols: int = 1,
) -> Path:
    """Write one styled sheet per table into a single .xlsx workbook.

    Parameters
    ----------
    sheets : dict of {str: pd.DataFrame}
        Mapping of sheet name to the table to write. Every table in the
        mapping must share the same header depth and index depth (i.e. come
        from the same table-building function), since a single style is
        applied to the whole workbook. Sheet names are truncated to Excel's
        31-character limit.
    file_path : str or Path
        Output file path. If it has no suffix, ".xlsx" is appended. The
        parent directory is created if needed.
    n_header_rows : int, default 1
        Number of header rows written by pandas for these tables: the column
        MultiIndex depth, plus one if that depth is greater than 1 (pandas
        adds one extra row for the row-index name in that case).
    n_index_cols : int, default 1
        Row-index depth (1 for a plain Index, 2+ for a MultiIndex).

    Returns
    -------
    Path
        Path where the workbook was saved.

    Raises
    ------
    ValueError
        If `sheets` is empty.
    RuntimeError
        If the workbook cannot be written due to a backend error.
    """
    if not sheets:
        raise ValueError("sheets must contain at least one table.")

    path = Path(file_path).expanduser()
    if path.suffix == "":
        path = path.with_suffix(".xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            for sheet_name, table in sheets.items():
                table.to_excel(
                    writer,
                    sheet_name=sheet_name[:31],
                    merge_cells=True,
                    na_rep=NA_LABEL,
                )

        workbook = load_workbook(path)
        for sheet_name, table in sheets.items():
            worksheet = workbook[sheet_name[:31]]
            _style_worksheet(
                worksheet,
                n_header_rows=n_header_rows,
                n_index_cols=n_index_cols,
                n_data_rows=table.shape[0],
            )
        workbook.save(path)
    except Exception as exc:  # pragma: no cover - backend errors vary
        raise RuntimeError(f"Unable to write Excel workbook: {path}") from exc

    return path
